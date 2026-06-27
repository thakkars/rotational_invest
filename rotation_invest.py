import os
import time
import json
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

# -------------------------------
# CONFIG
# -------------------------------
FILE_PATH = os.getenv("FILE_PATH", "output/rotation_invest.xlsx")
HTML_PATH = os.getenv("HTML_PATH", "output/index.html")

RAW_SHEET       = "rawdata"
GB_SIGNAL_SHEET = "GB_signal"
SB_SIGNAL_SHEET = "SB_signal"

LOOKBACK    = 20
MAX_RETRIES = 3
RETRY_SLEEP = 5

YF_SYMBOLS = {
    "NiftyBees":  "NIFTYBEES.NS",
    "GoldBees":   "GOLDBEES.NS",
    "SilverBees": "SILVERBEES.NS",
}

os.makedirs(os.path.dirname(FILE_PATH), exist_ok=True)

# -------------------------------
# HELPERS
# -------------------------------
def latest_raw_date(raw_df):
    temp = raw_df.copy()
    temp["Date"] = pd.to_datetime(temp["Date"], dayfirst=True, errors="coerce")
    temp = temp.dropna(subset=["Date"])
    return None if temp.empty else temp["Date"].max().normalize()

def fetch_yf_data(ticker):
    last_err = None
    for _ in range(MAX_RETRIES):
        try:
            df = yf.download(ticker, period="20y", interval="1d", auto_adjust=True, progress=False)
            if df is None or df.empty:
                raise ValueError(f"No data for {ticker}")
            df = df.reset_index()
            # yfinance may return MultiIndex columns — flatten
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = [c[0] for c in df.columns]
            df.rename(columns={"Date": "Date", "Open": "Open", "High": "High",
                                "Low": "Low", "Close": "Close", "Volume": "Volume"}, inplace=True)
            df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.normalize()
            df = df[["Date", "Open", "High", "Low", "Close", "Volume"]].copy()
            for c in ["Open", "High", "Low", "Close"]:
                df[c] = pd.to_numeric(df[c], errors="coerce").round(2)
            df = df.dropna(subset=["Date", "Close"])
            return df
        except Exception as e:
            last_err = e
            time.sleep(RETRY_SLEEP)
    raise ValueError(f"Failed to fetch {ticker} after {MAX_RETRIES} retries: {last_err}")

def fmt2(x):
    return round(float(x), 2) if pd.notna(x) else None

def build_signals(df, ratio_col, long_a, long_b, a_open, b_open, a_close, b_close, out_b_close):
    d = df[["Date", a_open, b_open, a_close, b_close, ratio_col]].copy()
    d = d.dropna(subset=[ratio_col]).reset_index(drop=True)
    d["20D_H_Ratio"] = d[ratio_col].rolling(LOOKBACK).max().shift(1)
    d["20D_L_Ratio"] = d[ratio_col].rolling(LOOKBACK).min().shift(1)
    signals = []
    holding = None
    for i, row in d.iterrows():
        if pd.isna(row["20D_H_Ratio"]) or pd.isna(row["20D_L_Ratio"]):
            continue
        if row[ratio_col] > row["20D_H_Ratio"] and holding != long_a:
            holding = long_a
            signals.append((i, row["Date"], f"BUY {long_a}"))
        elif row[ratio_col] < row["20D_L_Ratio"] and holding != long_b:
            holding = long_b
            signals.append((i, row["Date"], f"BUY {long_b}"))
    rows = []
    for k, (sig_idx, sig_date, buy_signal) in enumerate(signals):
        nxt = d[d.index > sig_idx]
        if not nxt.empty:
            entry_row   = nxt.iloc[0]
            entry_date  = entry_row["Date"]
            entry_price = entry_row[a_open] if buy_signal.endswith(long_a) else entry_row[b_open]
        else:
            entry_date  = pd.NaT
            entry_price = None
        if k + 1 < len(signals):
            ni  = signals[k + 1][0]
            nx2 = d[d.index > ni]
            if not nx2.empty:
                exit_row   = nx2.iloc[0]
                exit_date  = exit_row["Date"]
                exit_price = exit_row[a_open] if buy_signal.endswith(long_a) else exit_row[b_open]
            else:
                exit_date  = pd.NaT
                exit_price = None
        else:
            exit_date  = pd.NaT
            exit_price = None
        returns = None
        if entry_price is not None and exit_price is not None and entry_price != 0:
            returns = round(((exit_price - entry_price) / entry_price) * 100, 2)
        rows.append({
            "Date":         sig_date,
            "Buy Signal":   buy_signal,
            "Entry_Date":   entry_date  if pd.notna(entry_date)  else None,
            "Entry_Price":  fmt2(entry_price),
            "Exit_Date":    exit_date   if pd.notna(exit_date)   else None,
            "Exit_Price":   fmt2(exit_price),
            "Returns":      f"{returns}%" if returns is not None else None,
            "20D_H_Ratio":  round(float(d.loc[sig_idx, "20D_H_Ratio"]), 4),
            "20D_L_Ratio":  round(float(d.loc[sig_idx, "20D_L_Ratio"]), 4),
            "NB_Close":     fmt2(d.loc[sig_idx, a_close]),
            out_b_close:    fmt2(d.loc[sig_idx, b_close]),
        })
    return pd.DataFrame(rows)

def apply_sheet_formatting(ws, date_cols):
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column in date_cols and cell.value is not None:
                cell.number_format = "dd-mm-yy"
            elif isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for r in range(1, min(ws.max_row, 500) + 1):
            v = ws.cell(r, col).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = max(12, min(max_len + 2, 20))

def safe(v):
    return round(float(v), 4) if pd.notna(v) else None

# -------------------------------
# HTML DASHBOARD (clean light theme)
# -------------------------------
CSS = """
*{box-sizing:border-box}
body{margin:0;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Arial,sans-serif;background:#f1f5f9;color:#0f172a;-webkit-font-smoothing:antialiased}
.wrap{max-width:980px;margin:0 auto;padding:16px}
header{display:flex;align-items:center;justify-content:space-between;margin-bottom:16px;flex-wrap:wrap;gap:8px}
header h1{margin:0;font-size:1.45rem;font-weight:800;letter-spacing:-.01em}
.updated{font-size:.78rem;color:#64748b;background:#fff;padding:6px 12px;border-radius:999px;border:1px solid #e2e8f0}
section{margin-bottom:14px}
.grid{display:grid;gap:12px}
.two{grid-template-columns:1fr 1fr}
.three{grid-template-columns:repeat(3,1fr)}
.card{background:#fff;border:1px solid #e2e8f0;border-radius:16px;padding:16px;box-shadow:0 1px 3px rgba(0,0,0,.04)}
.pos-title{font-size:.7rem;text-transform:uppercase;letter-spacing:.05em;color:#94a3b8;font-weight:700}
.pos-asset{display:inline-block;margin-top:8px;padding:7px 14px;border-radius:10px;color:#fff;font-weight:800;font-size:1.02rem}
.pos-meta{margin-top:10px;display:flex;gap:14px;font-size:.8rem;color:#64748b;flex-wrap:wrap}
.pos-unreal{margin-top:10px;font-size:1.5rem;font-weight:800}
.pos-unreal small{font-size:.68rem;color:#94a3b8;font-weight:600;margin-left:4px}
.muted{color:#94a3b8}
.price-l{font-size:.7rem;text-transform:uppercase;color:#94a3b8;font-weight:700}
.price-v{font-size:1.55rem;font-weight:800;margin-top:4px}
.price-c{font-size:.84rem;font-weight:700;margin-top:2px}
.gauge-top{display:flex;justify-content:space-between;align-items:center;font-weight:700;font-size:.95rem}
.gauge-top .ratio{color:#2563eb}
.hint{font-size:.66rem;padding:3px 9px;border-radius:999px;background:#eff6ff;color:#2563eb;font-weight:700;text-transform:uppercase;letter-spacing:.03em}
.hint.breakdown{background:#fef2f2;color:#ef4444}
.gauge-track{position:relative;height:10px;background:linear-gradient(90deg,#fde68a 0%,#fef3c7 35%,#dbeafe 65%,#bfdbfe 100%);border-radius:999px;margin:14px 0 6px}
.gauge-mark{position:absolute;top:-5px;width:4px;height:20px;background:#0f172a;border-radius:3px;transform:translateX(-50%)}
.gauge-labels{display:flex;justify-content:space-between;font-size:.68rem;color:#94a3b8}
.chart-box{height:240px;position:relative}
.chart-box h3{margin:0 0 6px;font-size:.92rem}
.stat-title{font-size:.7rem;text-transform:uppercase;color:#94a3b8;font-weight:700;margin-bottom:12px}
.stat-grid{display:grid;grid-template-columns:repeat(5,1fr);gap:8px}
.stat-l{display:block;font-size:.62rem;color:#94a3b8;text-transform:uppercase}
.stat-v{display:block;font-size:1.05rem;font-weight:800;margin-top:3px}
table{width:100%;border-collapse:collapse;font-size:.78rem}
th{background:#f8fafc;color:#64748b;font-size:.64rem;text-transform:uppercase;padding:9px 6px;text-align:center}
td{padding:8px 6px;text-align:center;border-top:1px solid #f1f5f9}
.pill{display:inline-block;padding:3px 9px;border-radius:999px;color:#fff;font-size:.68rem;font-weight:700}
.table-scroll{overflow-x:auto}
.tbl-title{font-weight:800;font-size:.88rem;margin-bottom:8px;padding-bottom:8px;border-bottom:2px solid #f1f5f9}
.action-card h3{margin:0 0 4px;font-size:1rem}
.action-card p{margin:0 0 12px;font-size:.8rem;color:#64748b}
.run-btn{background:#2563eb;color:#fff;border:none;padding:11px 24px;border-radius:10px;font-size:.92rem;font-weight:700;cursor:pointer;transition:background .15s}
.run-btn:hover{background:#1d4ed8}
.run-btn:disabled{background:#94a3b8;cursor:not-allowed}
.run-box{margin-top:14px}
.run-track{height:8px;background:#e2e8f0;border-radius:999px;overflow:hidden}
#runBar{height:100%;width:0;background:#2563eb;transition:width .4s ease}
#runText{margin-top:8px;font-size:.8rem;color:#475569}
.run-log{display:inline-block;margin-top:10px;font-size:.8rem;color:#2563eb;text-decoration:none}
.run-err{display:block;margin-top:10px;font-size:.8rem;color:#ef4444}
.run-err a{color:#2563eb}
.excel-link{display:block;text-align:center;background:#fff;border:1px solid #e2e8f0;color:#0f172a;padding:14px;border-radius:12px;font-weight:700;text-decoration:none;margin-top:12px}
.excel-link:active{background:#f8fafc}
.hidden{display:none!important}
@media(max-width:640px){
  .two,.three{grid-template-columns:1fr}
  .stat-grid{grid-template-columns:repeat(3,1fr)}
  .chart-box{height:200px}
}
"""

JS = """
function makeChart(id,main,hi,lo,color){
  new Chart(document.getElementById(id),{
    type:'line',
    data:{labels:D.labels,datasets:[
      {label:'Ratio',data:main,borderColor:color,borderWidth:3,pointRadius:0,tension:.3,fill:false},
      {label:'20D High',data:hi,borderColor:'#16a34a',borderDash:[6,4],borderWidth:2,pointRadius:0,fill:false},
      {label:'20D Low',data:lo,borderColor:'#ef4444',borderDash:[6,4],borderWidth:2,pointRadius:0,fill:false}
    ]},
    options:{responsive:true,maintainAspectRatio:false,
      plugins:{legend:{labels:{boxWidth:12,font:{size:11}}}},
      scales:{x:{ticks:{maxTicksLimit:7,font:{size:10}},grid:{display:false}},
              y:{ticks:{font:{size:10}}}}
    }
  });
}
makeChart('c1',D.nbgb,D.h1,D.l1,'#2563eb');
makeChart('c2',D.nbsb,D.h2,D.l2,'#0f172a');

(function(){
  const btn=document.getElementById('runBtn');
  const box=document.getElementById('runBox');
  const bar=document.getElementById('runBar');
  const txt=document.getElementById('runText');
  const log=document.getElementById('runLog');
  const err=document.getElementById('runErr');
  let timer=null,trigTs=0;
  function setBar(p,s){bar.style.width=p+'%';txt.innerText=s;}
  btn.addEventListener('click',async function(){
    btn.disabled=true;box.classList.remove('hidden');
    err.classList.add('hidden');log.classList.add('hidden');
    setBar(8,'Triggering workflow…');
    trigTs=Date.now();
    try{
      const r=await fetch('/api/trigger',{method:'POST'});
      const d=await r.json();
      if(!d.ok)throw 0;
      setBar(15,'Workflow queued…');
      startPoll();
    }catch(e){
      setBar(0,'');
      err.classList.remove('hidden');
      err.innerHTML='Could not trigger automatically. <a href="https://github.com/thakkars/rotational_invest/actions" target="_blank">Run it manually on GitHub →</a>';
      btn.disabled=false;
    }
  });
  function startPoll(){
    timer=setInterval(async function(){
      try{
        const r=await fetch('/api/status');
        const d=await r.json();
        if(!d)return;
        let ca=d.created_at?new Date(d.created_at).getTime():0;
        if(ca && ca<trigTs-30000){setBar(15,'Queued…');return;}
        if(d.html_url){log.href=d.html_url;log.classList.remove('hidden');}
        if(d.status==='queued'){setBar(15,'Queued…');}
        else if(d.status==='in_progress'){
          let p=Math.min(90,25+Math.floor((Date.now()-trigTs)/900));
          setBar(p,'Running… fetching latest data');
        }else if(d.status==='completed'){
          clearInterval(timer);
          if(d.conclusion==='success'){setBar(100,'Done! Reloading…');setTimeout(function(){location.reload();},1200);}
          else{setBar(100,'Finished with issues — check logs');btn.disabled=false;}
        }
      }catch(e){}
    },4000);
  }
})();
"""

def generate_html(raw_df, gb_sig, sb_sig):

    # --- chart data (last 70 sessions) ---
    chart = raw_df[['Date','NB/GB','20D_H_NB_GB','20D_L_NB_GB',
                    'NB/SB','20D_H_NB_SB','20D_L_NB_SB',
                    'NB_close','GB_close','SB_close']].dropna(subset=['NB/GB']).tail(70).copy()
    chart['Date'] = chart['Date'].dt.strftime('%d-%m-%y')

    # --- latest snapshot ---
    lv = raw_df.dropna(subset=['20D_H_NB_GB','20D_L_NB_GB']).iloc[-1]
    latest_date = lv['Date'].strftime('%d-%m-%y')
    nb_price = round(float(lv['NB_close']),2)
    gb_price = round(float(lv['GB_close']),2)
    sb_price = round(float(lv['SB_close']),2)

    nb_gb  = round(float(lv['NB/GB']),4)
    nb_sb  = round(float(lv['NB/SB']),4)
    h_nbgb = round(float(lv['20D_H_NB_GB']),4)
    l_nbgb = round(float(lv['20D_L_NB_GB']),4)
    h_nbsb = round(float(lv['20D_H_NB_SB']),4)
    l_nbsb = round(float(lv['20D_L_NB_SB']),4)

    # --- today's % change vs previous session ---
    closes = raw_df.dropna(subset=['NB_close','GB_close','SB_close'])
    prev = closes.iloc[-2] if len(closes) >= 2 else lv
    def chg(cur, old): return round((cur-old)/old*100, 2) if old else 0.0
    nb_chg = chg(nb_price, prev['NB_close'])
    gb_chg = chg(gb_price, prev['GB_close'])
    sb_chg = chg(sb_price, prev['SB_close'])

    # --- ratio gauge positions (0-100) ---
    def gauge(v, hi, lo):
        if hi == lo: return 50.0
        return max(0.0, min(100.0, (v-lo)/(hi-lo)*100))
    gb_gauge = round(gauge(nb_gb, h_nbgb, l_nbgb),1)
    sb_gauge = round(gauge(nb_sb, h_nbsb, l_nbsb),1)

    # --- chart arrays ---
    dates = chart['Date'].tolist()
    nbgb_v = [round(float(x),4) for x in chart['NB/GB']]
    h1_v   = [round(float(x),4) for x in chart['20D_H_NB_GB']]
    l1_v   = [round(float(x),4) for x in chart['20D_L_NB_GB']]
    nbsb_v = [round(float(x),4) for x in chart['NB/SB']]
    h2_v   = [round(float(x),4) for x in chart['20D_H_NB_SB']]
    l2_v   = [round(float(x),4) for x in chart['20D_L_NB_SB']]

    data_json = json.dumps({
        'labels': dates,
        'nbgb': nbgb_v, 'h1': h1_v, 'l1': l1_v,
        'nbsb': nbsb_v, 'h2': h2_v, 'l2': l2_v,
    })

    price_map = {'NiftyBees':nb_price,'GoldBees':gb_price,'SilverBees':sb_price}
    ASSET_COLOR = {'NiftyBees':'#2563eb','GoldBees':'#d97706','SilverBees':'#64748b'}

    def fdate(d):
        if d is None or (not isinstance(d,str) and pd.isna(d)): return '—'
        try: return pd.to_datetime(d).strftime('%d-%m-%y')
        except Exception: return str(d)

    # --- per-strategy position + stats ---
    def strategy_info(sig):
        if sig is None or sig.empty: return None
        last = sig.iloc[-1]
        asset = str(last['Buy Signal']).replace('BUY ','')
        entry = last['Entry_Price']
        entry_date = last['Entry_Date']
        open_now = pd.isna(last['Exit_Price'])
        unreal = None
        if pd.notna(entry) and open_now and asset in price_map:
            unreal = round((price_map[asset]-float(entry))/float(entry)*100,2)
        rets = []
        for r in sig['Returns'].iloc[:-1]:
            if pd.notna(r):
                try: rets.append(float(str(r).replace('%','')))
                except Exception: pass
        win_rate = round(len([x for x in rets if x>0])/len(rets)*100,1) if rets else 0.0
        avg_ret  = round(sum(rets)/len(rets),2) if rets else 0.0
        best  = round(max(rets),2) if rets else 0.0
        worst = round(min(rets),2) if rets else 0.0
        return {'asset':asset,'entry':entry,'entry_date':entry_date,
                'unreal':unreal,'open':open_now,
                'total':len(sig),'closed':len(rets),
                'win_rate':win_rate,'avg':avg_ret,'best':best,'worst':worst}

    gb_info = strategy_info(gb_sig)
    sb_info = strategy_info(sb_sig)

    # --- position card ---
    def pos_card(title, info):
        if not info:
            return f'<div class="card"><div class="pos-title">{title}</div><div class="muted" style="margin-top:10px">No data</div></div>'
        a = info['asset']; col = ASSET_COLOR.get(a,'#0f172a')
        entry_s = f"{float(info['entry']):.2f}" if pd.notna(info['entry']) else '—'
        if info['open'] and info['unreal'] is not None:
            u = info['unreal']; ucol = '#16a34a' if u>=0 else '#ef4444'
            uhtml = f'<span style="color:{ucol}">{"+" if u>=0 else ""}{u}%</span><small>unrealized</small>'
        else:
            uhtml = '<span class="muted">—</span>'
        return f'''<div class="card">
          <div class="pos-title">{title}</div>
          <span class="pos-asset" style="background:{col}">{a}</span>
          <div class="pos-meta"><span>Since {fdate(info['entry_date'])}</span><span>Entry ₹{entry_s}</span></div>
          <div class="pos-unreal">{uhtml}</div>
        </div>'''

    # --- price card ---
    def price_card(label, price, chg, color):
        arrow = '▲' if chg>=0 else '▼'
        ccol = '#16a34a' if chg>=0 else '#ef4444'
        return f'''<div class="card">
          <div class="price-l">{label}</div>
          <div class="price-v" style="color:{color}">₹{price}</div>
          <div class="price-c" style="color:{ccol}">{arrow} {chg}%</div>
        </div>'''

    # --- gauge card ---
    def gauge_card(label, value, hi, lo, pct, low_asset):
        hint_cls = ''
        if pct >= 99: hint = 'AT BREAKOUT → Nifty'
        elif pct <= 1: hint = 'AT BREAKDOWN → ' + low_asset; hint_cls = ' breakdown'
        else: hint = 'in range'
        return f'''<div class="card">
          <div class="gauge-top"><span>{label} <span class="ratio">{value}</span></span><span class="hint{hint_cls}">{hint}</span></div>
          <div class="gauge-track"><div class="gauge-mark" style="left:{pct}%"></div></div>
          <div class="gauge-labels"><span>↓ {lo} (Hold {low_asset})</span><span>(Hold Nifty) {hi} ↑</span></div>
        </div>'''

    # --- stats card ---
    def stats_card(title, info):
        if not info: return ''
        best_html = f'<span style="color:#16a34a">+{info["best"]}%</span>' if info['best']>=0 else f'<span style="color:#ef4444">{info["best"]}%</span>'
        worst_html = f'<span style="color:#16a34a">+{info["worst"]}%</span>' if info['worst']>=0 else f'<span style="color:#ef4444">{info["worst"]}%</span>'
        return f'''<div class="card">
          <div class="stat-title">{title}</div>
          <div class="stat-grid">
            <div><span class="stat-l">Win Rate</span><span class="stat-v">{info["win_rate"]}%</span></div>
            <div><span class="stat-l">Trades</span><span class="stat-v">{info["closed"]}</span></div>
            <div><span class="stat-l">Avg</span><span class="stat-v">{info["avg"]}%</span></div>
            <div><span class="stat-l">Best</span><span class="stat-v">{best_html}</span></div>
            <div><span class="stat-l">Worst</span><span class="stat-v">{worst_html}</span></div>
          </div>
        </div>'''

    # --- trade history rows (last 8, newest first) ---
    def table_rows(sig):
        rows = ''
        for _, r in sig.tail(8).iloc[::-1].iterrows():
            color = ASSET_COLOR.get(str(r['Buy Signal']).replace('BUY ',''),'#0f172a')
            ret = r['Returns']
            if pd.isna(ret) or ret is None:
                ret_html = '<td class="muted">—</td>'
            else:
                val = str(ret); rcol = '#16a34a' if '-' not in val else '#ef4444'
                ret_html = f'<td style="color:{rcol};font-weight:700">{val}</td>'
            entry = f"{float(r['Entry_Price']):.2f}" if pd.notna(r['Entry_Price']) else '—'
            exitp = f"{float(r['Exit_Price']):.2f}" if pd.notna(r['Exit_Price']) else '—'
            rows += (
                f"<tr>"
                f"<td>{fdate(r['Date'])}</td>"
                f"<td><span class='pill' style='background:{color}'>{str(r['Buy Signal']).replace('BUY ','')}</span></td>"
                f"<td>{entry}</td>"
                f"<td>{exitp}</td>"
                f"{ret_html}"
                f"<td>{r['20D_H_Ratio']}</td>"
                f"<td>{r['20D_L_Ratio']}</td>"
                f"</tr>"
            )
        return rows

    def table_block(title, sig):
        return f'''<div class="card">
          <div class="tbl-title">{title}</div>
          <div class="table-scroll"><table>
            <thead><tr><th>Date</th><th>Signal</th><th>Entry</th><th>Exit</th><th>Returns</th><th>20D High</th><th>20D Low</th></tr></thead>
            <tbody>{table_rows(sig)}</tbody>
          </table></div>
        </div>'''

    gb_pos = pos_card('NB / GB Rotation', gb_info)
    sb_pos = pos_card('NB / SB Rotation', sb_info)
    gb_stats = stats_card('NB / GB Performance', gb_info)
    sb_stats = stats_card('NB / SB Performance', sb_info)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Rotation Strategy</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>{CSS}</style>
</head>
<body>
<div class="wrap">

  <header>
    <h1>Rotation Strategy</h1>
    <span class="updated">Updated {latest_date}</span>
  </header>

  <section class="grid two">
    {gb_pos}
    {sb_pos}
  </section>

  <section class="grid three">
    {price_card('NiftyBees', nb_price, nb_chg, '#2563eb')}
    {price_card('GoldBees', gb_price, gb_chg, '#d97706')}
    {price_card('SilverBees', sb_price, sb_chg, '#64748b')}
  </section>

  <section class="grid two">
    {gauge_card('NB / GB', nb_gb, h_nbgb, l_nbgb, gb_gauge, 'Gold')}
    {gauge_card('NB / SB', nb_sb, h_nbsb, l_nbsb, sb_gauge, 'Silver')}
  </section>

  <section class="grid two">
    <div class="card chart-box"><h3>NB / GB Ratio</h3><canvas id="c1"></canvas></div>
    <div class="card chart-box"><h3>NB / SB Ratio</h3><canvas id="c2"></canvas></div>
  </section>

  <section class="grid two">
    {gb_stats}
    {sb_stats}
  </section>

  <section class="grid two">
    {table_block('NB / GB Events', gb_sig)}
    {table_block('NB / SB Events', sb_sig)}
  </section>

  <section>
    <div class="card action-card">
      <h3>Manual Refresh</h3>
      <p>Force-update the dashboard with the latest market data.</p>
      <button id="runBtn" class="run-btn">▶ Run Now</button>
      <div id="runBox" class="run-box hidden">
        <div class="run-track"><div id="runBar"></div></div>
        <div id="runText">Starting…</div>
      </div>
      <a id="runLog" class="run-log hidden" href="#" target="_blank">View run on GitHub →</a>
      <span id="runErr" class="run-err hidden"></span>
    </div>
    <a class="excel-link" href="rotation_invest.xlsx" download>⬇ Download Excel (all signals)</a>
  </section>

</div>
<script>const D={data_json};</script>
<script>{JS}</script>
</body>
</html>"""

    return html

# -------------------------------
# LOAD or CREATE WORKBOOK
# -------------------------------
if os.path.exists(FILE_PATH):
    wb     = load_workbook(FILE_PATH)
    raw_df = pd.read_excel(FILE_PATH, sheet_name=RAW_SHEET)
else:
    wb     = Workbook()
    wb.create_sheet(RAW_SHEET, 0)
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    raw_df = pd.DataFrame()

today     = pd.Timestamp.today().normalize()
last_date = latest_raw_date(raw_df) if not raw_df.empty else None
needs_update = (last_date is None) or (last_date < today)

print(f"Today: {today.strftime('%d-%m-%y')}")
print(f"Latest raw data: {last_date.strftime('%d-%m-%y') if last_date is not None else 'None'}")
print(f"Needs update: {needs_update}")

# -------------------------------
# FETCH + UPDATE RAWDATA
# -------------------------------
try:
    nb = fetch_yf_data(YF_SYMBOLS["NiftyBees"]).rename(columns={
        "Open":"NB_open","High":"NB_high","Low":"NB_low","Close":"NB_close","Volume":"NB_volume"
    })[["Date","NB_open","NB_high","NB_low","NB_close","NB_volume"]]
    gb = fetch_yf_data(YF_SYMBOLS["GoldBees"]).rename(columns={
        "Open":"GB_open","High":"GB_high","Low":"GB_low","Close":"GB_close","Volume":"GB_volume"
    })[["Date","GB_open","GB_high","GB_low","GB_close","GB_volume"]]
    sb = fetch_yf_data(YF_SYMBOLS["SilverBees"]).rename(columns={
        "Open":"SB_open","High":"SB_high","Low":"SB_low","Close":"SB_close","Volume":"SB_volume"
    })[["Date","SB_open","SB_high","SB_low","SB_close","SB_volume"]]
    raw_updated = nb.merge(gb, on="Date", how="outer").merge(sb, on="Date", how="outer")
    raw_updated = raw_updated.sort_values("Date").reset_index(drop=True)
    raw_updated["NB/GB"] = (raw_updated["NB_close"] / raw_updated["GB_close"]).round(4)
    raw_updated["NB/SB"] = (raw_updated["NB_close"] / raw_updated["SB_close"]).round(4)
    if RAW_SHEET in wb.sheetnames:
        del wb[RAW_SHEET]
    ws_raw = wb.create_sheet(RAW_SHEET, 0)
    for c_idx, h in enumerate(raw_updated.columns, 1):
        ws_raw.cell(1, c_idx, h)
    for r_idx, row in enumerate(raw_updated.itertuples(index=False), 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws_raw.cell(r_idx, c_idx)
            if c_idx == 1 and pd.notna(val):
                cell.value = pd.to_datetime(val).to_pydatetime()
                cell.number_format = "dd-mm-yy"
            elif isinstance(val, (int, float)) and pd.notna(val):
                if c_idx in [2,3,4,5,7,8,9,10,12,13,14,15]:
                    cell.value = round(float(val), 2)
                    cell.number_format = "0.00"
                elif c_idx in [17,18]:
                    cell.value = round(float(val), 4)
                    cell.number_format = "0.0000"
                else:
                    cell.value = val
            else:
                cell.value = val
    apply_sheet_formatting(ws_raw, date_cols={1})
    print("Raw data fetched and updated successfully.")
except Exception as e:
    print(f"Fetch failed: {e}")
    if raw_df.empty:
        raise RuntimeError("No existing data and fresh fetch failed — cannot proceed.") from e
    raw_updated = raw_df.copy()
    print("Using existing raw data from workbook.")

raw_updated["Date"] = pd.to_datetime(raw_updated["Date"], dayfirst=True, errors="coerce")
raw_updated = raw_updated.sort_values("Date").reset_index(drop=True)

raw_updated["NB/GB"] = raw_updated["NB_close"] / raw_updated["GB_close"]
raw_updated["NB/SB"] = raw_updated["NB_close"] / raw_updated["SB_close"]
raw_updated["20D_H_NB_GB"] = raw_updated["NB/GB"].rolling(20).max().shift(1)
raw_updated["20D_L_NB_GB"] = raw_updated["NB/GB"].rolling(20).min().shift(1)
raw_updated["20D_H_NB_SB"] = raw_updated["NB/SB"].rolling(20).max().shift(1)
raw_updated["20D_L_NB_SB"] = raw_updated["NB/SB"].rolling(20).min().shift(1)

gb_signals = build_signals(raw_updated,"NB/GB","NiftyBees","GoldBees","NB_open","GB_open","NB_close","GB_close","GB_Close")
sb_signals = build_signals(raw_updated,"NB/SB","NiftyBees","SilverBees","NB_open","SB_open","NB_close","SB_close","SB_Close")

for sheet_name, sig_df in [(GB_SIGNAL_SHEET, gb_signals),(SB_SIGNAL_SHEET, sb_signals)]:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    for c_idx, h in enumerate(sig_df.columns, 1):
        ws.cell(1, c_idx, h)
    for r_idx, row in enumerate(sig_df.itertuples(index=False), 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx)
            if c_idx in [1,3,5] and val is not None:
                try:
                    cell.value = pd.to_datetime(val).to_pydatetime()
                    cell.number_format = "dd-mm-yy"
                except:
                    cell.value = val
            elif isinstance(val, str) and val.endswith("%"):
                cell.value = val
            elif isinstance(val, (int, float)) and pd.notna(val):
                if c_idx in [4,6,10,11]:
                    cell.value = round(float(val), 2)
                    cell.number_format = "0.00"
                elif c_idx in [8,9]:
                    cell.value = round(float(val), 4)
                    cell.number_format = "0.0000"
                else:
                    cell.value = round(float(val), 2)
                    cell.number_format = "0.00"
            else:
                cell.value = val
    apply_sheet_formatting(ws, date_cols={1,3,5})

if RAW_SHEET in wb.sheetnames:
    wb._sheets = [wb[RAW_SHEET]] + [wb[s] for s in wb.sheetnames if s != RAW_SHEET]

wb.save(FILE_PATH)
print(f"Workbook saved: {FILE_PATH}")
print(f"GB signals: {len(gb_signals)}")
print(f"SB signals: {len(sb_signals)}")

html_content = generate_html(raw_updated, gb_signals, sb_signals)
with open(HTML_PATH, 'w', encoding='utf-8') as f:
    f.write(html_content)
print(f"Dashboard saved: {HTML_PATH}")
